"""doaya-sayfa-satırlar-sütunlar-hücre
workbook oluştururuz-sheet-row/column-cell-value"""

import openpyxl
dosya=openpyxl.load_workbook("./data.xlsx") #buraya dosyanın bulunduğu konumu yazarız.aynı dosyada ise bunu yapmamız yeterli değilse C: yapıp konumunu yazarız.
print(dosya.sheetnames)#sayfa isimlerini bize liste olarak verir
print("aktif sayfa:"+ dosya.active.title)# bu komutla da aktif sayfanın adını yazdırırız.
sayfa= dosya["kitaplar"] #kitaplar sayfasını biz sayfa olarak kaydettik
deger=sayfa["B4"].value# bu sayfanın içindeki b4 hücresini verir.
print(deger)
"""tek seferde bir hücrenin verisini almak için 
deger=dosya["kitaplar]["B4"].value
"""
deger= dosya["kitaplar"]
veri=sayfa.cell(3,2).value #belirtilen sayfadaki 3. satır 2. sütunun içindeki veriyi verir.
print(veri)
