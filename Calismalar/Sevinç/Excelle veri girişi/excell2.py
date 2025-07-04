import openpyxl


def excel_listeler_listesine_cevir(dosyayer, sayfaadi):
    dosya = openpyxl.load_workbook(dosyayer)
    sayfa = dosya[sayfaadi]
    satir_sayisi = sayfa.max_row
    sutun_sayisi = sayfa.max_column
    data = []
    for i in range(1, satir_sayisi + 1):
        satir = []
        for j in range(1, sutun_sayisi + 1):
            satir.append(sayfa.cell(i, j).value)
        data.append(satir)
    return data


print(excel_listeler_listesine_cevir("./data.xlsx", "notlar"))
